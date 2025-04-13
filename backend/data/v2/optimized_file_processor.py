import os
import logging
import traceback
import numpy as np
import pandas as pd
import chardet
from datetime import datetime
import multiprocessing
from functools import partial
import concurrent.futures
import gc
import psutil
import time
from io import StringIO, BytesIO
import threading
import queue

from ..file_processor import FileProcessor, FileTypeDetector, handle_nan_values, generate_column_info

logger = logging.getLogger(__name__)


class OptimizedFileProcessor:
    """
    Optimized version of FileProcessor that maximizes CPU and memory utilization
    for faster data processing and extraction.
    """
    # Number of worker processes to use for parallel processing
    # By default use N-1 cores, leaving one for system processes
    DEFAULT_WORKERS = max(1, multiprocessing.cpu_count() - 1)

    # Chunk size for processing large files
    DEFAULT_CHUNK_SIZE = 10000

    # Memory threshold (70% of system memory)
    MEMORY_THRESHOLD = 0.7

    @staticmethod
    def get_available_memory():
        """Get available system memory in bytes"""
        return psutil.virtual_memory().available

    @staticmethod
    def get_optimal_chunk_size(file_size, available_memory):
        """Calculate optimal chunk size based on file size and available memory"""
        # Estimate that processed data will be ~3x the size of raw data
        processing_factor = 3.0

        # Target using 60% of available memory
        memory_to_use = available_memory * 0.6

        # Calculate rows per chunk
        estimated_row_size = 2000  # Rough estimate of bytes per row
        optimal_chunk_size = int(
            memory_to_use / (estimated_row_size * processing_factor))

        # Cap at reasonable values
        return max(1000, min(optimal_chunk_size, 50000))

    @staticmethod
    def process_file(file_path, file_name, max_workers=None, chunk_size=None):
        """
        Process a file with parallel processing for improved performance

        Args:
            file_path (str): Path to the file to process
            file_name (str): Name of the file
            max_workers (int, optional): Maximum number of worker processes
            chunk_size (int, optional): Size of chunks to process

        Returns:
            tuple: (processed_data, summary_data)
        """
        start_time = time.time()

        try:
            logger.info(
                f"OptimizedFileProcessor: Processing file: {file_name}")

            # Check if file exists
            if not os.path.exists(file_path):
                logger.error(f"File not found: {file_path}")
                return {"error": "File not found"}, {"error": "File not found"}

            # Get file size
            file_size = os.path.getsize(file_path)
            logger.info(f"File size: {file_size / (1024*1024):.2f} MB")

            # Calculate optimal parameters if not provided
            available_memory = OptimizedFileProcessor.get_available_memory()
            if not max_workers:
                max_workers = OptimizedFileProcessor.DEFAULT_WORKERS

            if not chunk_size and file_size > 10 * 1024 * 1024:  # If file > 10MB
                chunk_size = OptimizedFileProcessor.get_optimal_chunk_size(
                    file_size, available_memory)
            elif not chunk_size:
                chunk_size = OptimizedFileProcessor.DEFAULT_CHUNK_SIZE

            logger.info(
                f"Using {max_workers} workers and chunk size of {chunk_size}")

            # Detect file type
            detector = FileTypeDetector()
            file_type, confidence, algorithm = detector.detect_file_type(
                file_path, file_name)

            logger.info(
                f"Detected file type: {file_type} with confidence {confidence}, using algorithm: {algorithm}")

            # Get the specific processing method for this file type
            if hasattr(FileProcessor, algorithm):
                base_processing_method = getattr(FileProcessor, algorithm)
            else:
                logger.warning(
                    f"Processing method {algorithm} not found, using generic processor")
                base_processing_method = FileProcessor.process_generic

            # For small files, just use the original processor
            if file_size < 5 * 1024 * 1024:  # Less than 5MB
                logger.info(f"File size is small, using original processor")
                result, summary = base_processing_method(file_path)

                # Add detected file type to summary
                if summary and isinstance(summary, dict):
                    summary['detected_file_type'] = file_type
                    summary['detection_confidence'] = handle_nan_values(
                        confidence)
                    summary['processing_time_seconds'] = time.time() - \
                        start_time

                return result, summary

            # For larger files, use our optimized approach
            try:
                # For Excel files
                if file_path.endswith(('.xlsx', '.xls')):
                    result, summary = OptimizedFileProcessor._process_excel_file(
                        file_path, base_processing_method, max_workers, chunk_size
                    )
                # For CSV files
                elif file_path.endswith('.csv'):
                    result, summary = OptimizedFileProcessor._process_csv_file(
                        file_path, base_processing_method, max_workers, chunk_size
                    )
                # For other file types
                else:
                    # Fall back to original processor
                    logger.info(
                        f"Unsupported file type for optimization, using original processor")
                    result, summary = base_processing_method(file_path)

                # Add detected file type to summary
                if summary and isinstance(summary, dict):
                    summary['detected_file_type'] = file_type
                    summary['detection_confidence'] = handle_nan_values(
                        confidence)
                    summary['processing_time_seconds'] = time.time() - \
                        start_time
                    summary['optimized_processing'] = True
                    summary['workers_used'] = max_workers
                    summary['chunk_size_used'] = chunk_size

                return result, summary

            except Exception as e:
                logger.error(f"Error in optimized processing: {str(e)}")
                logger.error(traceback.format_exc())
                logger.info(f"Falling back to original processor")

                # Fall back to original processor
                result, summary = base_processing_method(file_path)

                # Add detected file type to summary
                if summary and isinstance(summary, dict):
                    summary['detected_file_type'] = file_type
                    summary['detection_confidence'] = handle_nan_values(
                        confidence)
                    summary['processing_time_seconds'] = time.time() - \
                        start_time
                    summary['optimized_processing'] = False
                    summary['fallback_reason'] = str(e)

                return result, summary

        except Exception as e:
            logger.error(f"Error in optimized process_file: {str(e)}")
            logger.error(traceback.format_exc())
            end_time = time.time()
            return {"error": str(e)}, {
                "error": str(e),
                "processing_time_seconds": end_time - start_time
            }

    @staticmethod
    def _process_excel_file(file_path, processing_method, max_workers, chunk_size):
        """Process an Excel file in parallel chunks"""
        logger.info(f"Processing Excel file in chunks: {file_path}")

        # First determine the total number of rows
        xl = pd.ExcelFile(file_path)
        sheet_name = xl.sheet_names[0]  # Use first sheet by default

        # Read the excel file headers only
        df_headers = pd.read_excel(xl, sheet_name=sheet_name, nrows=0)

        # Use specific Excel reader to get row count without loading whole file
        from openpyxl import load_workbook
        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb[sheet_name]
        # Get approximate row count
        row_count = ws.max_row
        wb.close()

        logger.info(f"Excel file has approximately {row_count} rows")

        # If small number of rows, just process directly
        if row_count <= chunk_size:
            df = pd.read_excel(xl, sheet_name=sheet_name)
            data = df.to_dict('records')

            # Run through the regular processor to get proper formatting
            # We're just passing the data dict, not the file path
            return OptimizedFileProcessor._process_dataframe(df, processing_method.__name__)

        # Process in chunks using parallel workers
        chunks = []
        for i in range(0, row_count, chunk_size):
            end_row = min(i + chunk_size, row_count)
            chunks.append((i, chunk_size))
            logger.debug(f"Added chunk: rows {i} to {end_row}")

        # Create a thread-safe queue for results
        result_queue = queue.Queue()

        # Define the worker function that processes a chunk
        def process_chunk(chunk_start, chunk_size):
            try:
                # Read the chunk from Excel
                df_chunk = pd.read_excel(
                    file_path, skiprows=chunk_start, nrows=chunk_size)

                # Process the dataframe using our helper
                chunk_result, _ = OptimizedFileProcessor._process_dataframe(
                    df_chunk, processing_method.__name__)

                # Put the result in the queue
                result_queue.put(chunk_result)

                # Force garbage collection to free memory
                del df_chunk
                gc.collect()

                return True
            except Exception as e:
                logger.error(f"Error processing chunk {chunk_start}: {str(e)}")
                logger.error(traceback.format_exc())
                return False

        # Process chunks in parallel
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(process_chunk, start, size)
                       for start, size in chunks]
            concurrent.futures.wait(futures)

        # Combine all results
        all_results = []
        while not result_queue.empty():
            chunk_result = result_queue.get()
            if isinstance(chunk_result, list):
                all_results.extend(chunk_result)
            else:
                logger.warning(
                    f"Unexpected chunk result type: {type(chunk_result)}")

        # Generate summary for the combined results
        row_count = len(all_results)

        # Create a sample DataFrame for summary generation
        if all_results:
            sample_df = pd.DataFrame(all_results[:min(1000, len(all_results))])
            summary = {
                "row_count": row_count,
                "column_count": len(sample_df.columns) if not sample_df.empty else 0,
                "columns": generate_column_info(sample_df),
                "parallel_processing": True,
                "chunks_processed": len(chunks)
            }
        else:
            summary = {
                "row_count": 0,
                "column_count": 0,
                "columns": [],
                "error": "No results generated from chunks"
            }

        return all_results, summary

    @staticmethod
    def _process_csv_file(file_path, processing_method, max_workers, chunk_size):
        """Process a CSV file in parallel chunks"""
        logger.info(f"Processing CSV file in chunks: {file_path}")

        # First determine the encoding
        encodings_to_try = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']

        # Try different encodings
        for encoding in encodings_to_try:
            try:
                # Just check if we can read a small sample
                with open(file_path, 'r', encoding=encoding) as f:
                    f.read(1024)
                    break
            except UnicodeDecodeError:
                continue
        else:
            # If no encoding works, try detecting with chardet
            with open(file_path, 'rb') as file:
                raw_data = file.read(10000)  # Read a sample to detect encoding
                detected_encoding = chardet.detect(raw_data)['encoding']
                encoding = detected_encoding

        logger.info(f"Using encoding: {encoding}")

        # Determine the delimiter
        with open(file_path, 'r', encoding=encoding) as f:
            sample = f.read(1024)
            if sample.count(';') > sample.count(','):
                delimiter = ';'
            else:
                delimiter = ','

        logger.info(f"Using delimiter: {delimiter}")

        # Count total lines in file for better chunking
        with open(file_path, 'r', encoding=encoding) as f:
            row_count = sum(1 for _ in f) - 1  # Subtract header

        logger.info(f"CSV file has approximately {row_count} rows")

        # If small number of rows, just process directly
        if row_count <= chunk_size:
            df = pd.read_csv(file_path, delimiter=delimiter, encoding=encoding)
            return OptimizedFileProcessor._process_dataframe(df, processing_method.__name__)

        # Read the headers
        df_headers = pd.read_csv(
            file_path, delimiter=delimiter, encoding=encoding, nrows=0)
        headers = df_headers.columns.tolist()

        # Process in chunks using parallel workers
        chunks = []
        for i in range(0, row_count, chunk_size):
            end_row = min(i + chunk_size, row_count)
            # Skip 0 for first chunk because it includes header
            skip_rows = i if i == 0 else i + 1
            chunks.append((skip_rows, chunk_size))
            logger.debug(f"Added chunk: rows {i} to {end_row}")

        # Create a thread-safe queue for results
        result_queue = queue.Queue()

        # Define the worker function that processes a chunk
        def process_chunk(chunk_start, chunk_size):
            try:
                # Skip header except for first chunk
                skip_header = chunk_start != 0
                header = None if not skip_header else 0

                # Read the chunk from CSV
                df_chunk = pd.read_csv(
                    file_path,
                    delimiter=delimiter,
                    encoding=encoding,
                    skiprows=chunk_start,
                    nrows=chunk_size,
                    header=header,
                    dtype=object,  # Handle mixed data types as strings
                    low_memory=False  # Prevent dtype warnings
                )

                # If this isn't the first chunk, we need to add headers
                if chunk_start != 0:
                    df_chunk.columns = headers

                # Process the dataframe using our helper
                chunk_result, _ = OptimizedFileProcessor._process_dataframe(
                    df_chunk, processing_method.__name__)

                # Put the result in the queue
                result_queue.put(chunk_result)

                # Force garbage collection to free memory
                del df_chunk
                gc.collect()

                return True
            except Exception as e:
                logger.error(f"Error processing chunk {chunk_start}: {str(e)}")
                logger.error(traceback.format_exc())
                return False

        # Process chunks in parallel
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(process_chunk, start, size)
                       for start, size in chunks]
            concurrent.futures.wait(futures)

        # Combine all results
        all_results = []
        while not result_queue.empty():
            chunk_result = result_queue.get()
            if isinstance(chunk_result, list):
                all_results.extend(chunk_result)
            else:
                logger.warning(
                    f"Unexpected chunk result type: {type(chunk_result)}")

        # Generate summary for the combined results
        row_count = len(all_results)

        # Create a sample DataFrame for summary generation
        if all_results:
            sample_df = pd.DataFrame(all_results[:min(1000, len(all_results))])
            summary = {
                "row_count": row_count,
                "column_count": len(sample_df.columns) if not sample_df.empty else 0,
                "columns": generate_column_info(sample_df),
                "parallel_processing": True,
                "chunks_processed": len(chunks)
            }
        else:
            summary = {
                "row_count": 0,
                "column_count": 0,
                "columns": [],
                "error": "No results generated from chunks"
            }

        return all_results, summary

    @staticmethod
    def _process_dataframe(df, method_name):
        """
        Process a dataframe using the specified algorithm
        Args:
            df: pandas DataFrame
            method_name: name of the processing method
        Returns:
            tuple: (processed_data, summary)
        """
        try:
            # Create a dictionary of records
            records = df.to_dict('records')

            # Generate summary for the data
            column_info = generate_column_info(df)

            # Create summary based on common fields found in all methods
            summary = {
                "row_count": len(records),
                "column_count": len(df.columns),
                "columns": column_info
            }

            # Add method-specific summary items
            if 'facturation_manuelle' in method_name:
                if 'MONTANT_HT' in df.columns:
                    summary["total_ht"] = float(df['MONTANT_HT'].sum())
                if 'MONTANT_TTC' in df.columns:
                    summary["total_ttc"] = float(df['MONTANT_TTC'].sum())

            elif 'journal_ventes' in method_name:
                if 'CHIFFRE_AFF_EXE_DZD' in df.columns:
                    summary["total_revenue"] = float(
                        df['CHIFFRE_AFF_EXE_DZD'].sum())

            elif 'ca_periodique' in method_name or 'ca_non_periodique' in method_name:
                if 'HT' in df.columns:
                    summary["total_ht"] = float(df['HT'].sum())
                if 'TTC' in df.columns:
                    summary["total_ttc"] = float(df['TTC'].sum())
                if 'TAX' in df.columns:
                    summary["total_tax"] = float(df['TAX'].sum())

            # For DNT, RFD, CNT
            elif any(x in method_name for x in ['ca_dnt', 'ca_rfd', 'ca_cnt']):
                if 'HT' in df.columns:
                    summary["total_ht"] = float(df['HT'].sum())
                if 'TTC' in df.columns:
                    summary["total_ttc"] = float(df['TTC'].sum())
                if 'TVA' in df.columns:
                    summary["total_tva"] = float(df['TVA'].sum())

            # For etat_facture
            elif 'etat_facture' in method_name:
                if 'MONTANT_HT' in df.columns:
                    summary["total_ht"] = float(df['MONTANT_HT'].sum())
                if 'MONTANT_TTC' in df.columns:
                    summary["total_ttc"] = float(df['MONTANT_TTC'].sum())
                if 'MONTANT_TAXE' in df.columns:
                    summary["total_taxe"] = float(df['MONTANT_TAXE'].sum())
                if 'CHIFFRE_AFF_EXE' in df.columns:
                    summary["total_revenue"] = float(
                        df['CHIFFRE_AFF_EXE'].sum())

            # For creances_ngbss
            elif 'creances_ngbss' in method_name:
                if 'INVOICE_AMT' in df.columns:
                    summary["total_invoice_amt"] = float(
                        df['INVOICE_AMT'].sum())
                if 'OPEN_AMT' in df.columns:
                    summary["total_open_amt"] = float(df['OPEN_AMT'].sum())
                if 'CREANCE_NET' in df.columns:
                    summary["total_creance_net"] = float(
                        df['CREANCE_NET'].sum())

            return records, summary

        except Exception as e:
            logger.error(f"Error processing dataframe: {str(e)}")
            logger.error(traceback.format_exc())
            return [], {"error": str(e)}

    @staticmethod
    def bulk_process_files(file_paths, file_names, max_workers=None):
        """
        Process multiple files in parallel

        Args:
            file_paths (list): List of file paths
            file_names (list): List of file names
            max_workers (int, optional): Maximum number of worker processes

        Returns:
            dict: Dictionary of processed results by file path
        """
        if not max_workers:
            max_workers = min(
                len(file_paths), OptimizedFileProcessor.DEFAULT_WORKERS)

        logger.info(
            f"Bulk processing {len(file_paths)} files with {max_workers} workers")

        results = {}

        with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
            # Map file paths to futures
            future_to_path = {
                executor.submit(OptimizedFileProcessor.process_file, path, name): (path, name)
                for path, name in zip(file_paths, file_names)
            }

            # Process results as they complete
            for future in concurrent.futures.as_completed(future_to_path):
                path, name = future_to_path[future]
                try:
                    result, summary = future.result()
                    results[path] = {
                        'result': result,
                        'summary': summary,
                        'file_name': name
                    }
                except Exception as e:
                    logger.error(f"Error processing file {name}: {str(e)}")
                    logger.error(traceback.format_exc())
                    results[path] = {
                        'error': str(e),
                        'file_name': name
                    }

        return results

# Add compatibility with original FileProcessor interface


def process_file(file_path, file_name):
    """
    Drop-in replacement for FileProcessor.process_file 
    """
    return OptimizedFileProcessor.process_file(file_path, file_name)
